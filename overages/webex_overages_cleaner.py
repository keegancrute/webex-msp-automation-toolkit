#!/usr/bin/env python3

"""
Full Webex Overages Automation Pipeline

This script performs an end-to-end workflow for Webex license overage analysis:
    1. Cleans raw Webex overages CSV files
    2. Activates customer organizations in the Webex Partner API
    3. Retrieves all license data for each organization
    4. Computes overages and utilization
    5. Generates an Excel report with color-coded status

Outputs include:
    - cleaned_overages_<timestamp>.csv
    - successful_orgs_<timestamp>.json
    - webex_org_details_<timestamp>.csv
    - Webex_License_Overages_<timestamp>.xlsx

A Webex OAuth access token must be supplied via the WEBEX_ACCESS_TOKEN
environment variable.

Usage:
    export WEBEX_ACCESS_TOKEN="your_access_token_here"
    python3 webex_overages_cleaner.py
"""

import os
import sys
import csv
import json
import time
import requests
import pandas as pd
import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

# Constants
BASE_URL = "https://webexapis.com/v1"

# Load token from environment variable instead of hardcoding
ACCESS_TOKEN = os.environ.get("WEBEX_ACCESS_TOKEN")

if not ACCESS_TOKEN:
    print("ERROR: WEBEX_ACCESS_TOKEN environment variable is not set.")
    print('       export WEBEX_ACCESS_TOKEN="your_access_token_here"')
    sys.exit(1)

HEADERS = {
    "Authorization": f"Bearer {ACCESS_TOKEN}",
    "Content-Type": "application/json",
    "Accept": "application/json"
}

MAX_RETRIES = 3

def clean_csv(input_filename):
    today = datetime.datetime.today().strftime("%B_%d_%H-%M")
    output_filename = f"cleaned_overages_{today}.csv"

    with open(input_filename, "r", encoding="utf-8") as infile, open(output_filename, "w", encoding="utf-8", newline="") as outfile:
        reader = csv.reader(infile, delimiter=",", skipinitialspace=True)
        writer = csv.writer(outfile, quotechar='"', delimiter=",", quoting=csv.QUOTE_MINIMAL)

        for row in reader:
            if len(row) < 2:
                continue
            fixed_row = fix_misaligned_row(row)
            cleaned_row = fixed_row[:2]
            writer.writerow(cleaned_row)

    return output_filename

def fix_misaligned_row(row):
    if len(row) > 6:
        customer_name = " ".join(row[:-5])
        org_id = row[-5]
        return [customer_name, org_id]
    elif len(row) < 6:
        return row + [""] * (6 - len(row))
    return row

def load_cleaned_csv(filename):
    df = pd.read_csv(filename, quotechar='"')
    df = df.iloc[:, :2]
    df.columns = ["Customer Name", "Customer Org ID"]
    df = df.drop_duplicates()
    return df

def activate_organization(org_id):
    url = f"{BASE_URL}/organizations/{org_id}"
    response = requests.get(url, headers=HEADERS)
    return response.json() if response.status_code == 200 else None

def fetch_licenses(org_id):
    url = f"{BASE_URL}/licenses?orgId={org_id}"
    for attempt in range(MAX_RETRIES):
        response = requests.get(url, headers=HEADERS)
        if response.status_code == 403:
            return None
        if response.status_code == 429:
            time.sleep(10)
            continue
        try:
            response.raise_for_status()
            return response.json()
        except:
            time.sleep(3)
    return None

def gather_licenses(df):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    success_json = f"successful_orgs_{timestamp}.json"
    org_details_csv = f"webex_org_details_{timestamp}.csv"

    successful_orgs = {}
    org_details_list = []

    for _, row in df.iterrows():
        customer_name = row['Customer Name']
        org_id = row['Customer Org ID']
        print(f"Processing: {customer_name} ({org_id})")

        org_info = activate_organization(org_id)
        if not org_info:
            continue

        licenses = fetch_licenses(org_id)
        if licenses and "items" in licenses:
            successful_orgs[org_id] = {
                "customer_name": customer_name,
                "org_id": org_id,
                "org_display_name": org_info.get("displayName", "Unknown"),
                "created": org_info.get("created", "Unknown"),
                "country_code": org_info.get("countryCode", "Unknown"),
                "licenses": licenses["items"]
            }
            org_details_list.append({
                "Customer Name": customer_name,
                "Org ID": org_id,
                "Org Display Name": org_info.get("displayName", "Unknown"),
                "Created": org_info.get("created", "Unknown"),
                "Country Code": org_info.get("countryCode", "Unknown"),
                "License Count": len(licenses["items"])
            })
        time.sleep(1)

    with open(success_json, "w") as f:
        json.dump(successful_orgs, f, indent=4)
    pd.DataFrame(org_details_list).to_csv(org_details_csv, index=False)

    return success_json

def process_license_data(json_file):
    with open(json_file, "r") as f:
        data = json.load(f)

    report_data = []
    for org_id, org_info in data.items():
        org_name = org_info.get("customer_name", "Unknown Org")
        for license in org_info.get("licenses", []):
            name = license["name"]
            total_units = license["totalUnits"]
            consumed_units = license["consumedUnits"]
            overage = consumed_units - total_units
            status = "Overused" if overage > 0 else "Underutilized" if overage < 0 else "Fully Used"
            report_data.append({
                "Org Name": org_name,
                "License Name": name,
                "Total Units": total_units,
                "Consumed Units": consumed_units,
                "Status": status,
                "Overages": f"+{overage}" if overage > 0 else f"{overage}"
            })

    df = pd.DataFrame(report_data)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = f"Webex_License_Overages_{timestamp}.xlsx"
    df.to_excel(output_file, sheet_name="Overages", index=False)

    wb = load_workbook(output_file)
    ws = wb["Overages"]

    header_font = Font(bold=True)
    overage_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    underutilized_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        status_cell = row[4]
        for cell in row:
            if row[0].row == 2:
                cell.font = header_font
        if "Overused" in status_cell.value:
            for cell in row:
                cell.fill = overage_fill
        elif "Underutilized" in status_cell.value:
            for cell in row:
                cell.fill = underutilized_fill

    wb.save(output_file)
    print(f"✅ Final Excel report saved: {output_file}")

# --- Entry Point ---
if __name__ == "__main__":
    raw_path = input("Enter path to raw overages CSV: ").strip()
    cleaned_csv = clean_csv(raw_path)
    df = load_cleaned_csv(cleaned_csv)
    json_path = gather_licenses(df)
    process_license_data(json_path)
